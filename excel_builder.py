# ─────────────────────────────────────────────────────────────────────────────
# Generación del archivo Excel 
# ─────────────────────────────────────────────────────────────────────────────

from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config import ALL_COLS, SECTION_COLORS, COLUMN_WIDTHS, DEFAULT_WIDTH, TARIFAS, DF_OTRO_COLS


def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10): return Font(name="Arial", bold=bold, color=color, size=size)
def _aln(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


# ── Hoja CONCILIACION ─────────────────────────────────────────────────────────

def _write_conciliacion(ws, df: pd.DataFrame):
    bounds = {}
    for ci, (_, _, sec, _) in enumerate(ALL_COLS, 1):
        bounds.setdefault(sec, [ci, ci])
        bounds[sec][1] = ci

    for sec, (s, e) in bounds.items():
        c = ws.cell(1, s, sec)
        c.font      = _font(True, SECTION_COLORS[sec]["hfg"], 11)
        c.fill      = _fill(SECTION_COLORS[sec]["hbg"])
        c.alignment = _aln()
        if s != e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)

    for ci, (_, label, sec, _) in enumerate(ALL_COLS, 1):
        c = ws.cell(2, ci, label)
        c.font      = _font(True, "000000", 9)
        c.fill      = _fill(SECTION_COLORS[sec]["dbg"])
        c.alignment = _aln(wrap=True)
    ws.row_dimensions[2].height = 34

    # Datos
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        odd = (ri % 2 == 0)
        for ci, (field, _, sec, is_calc) in enumerate(ALL_COLS, 1):

            if field in df.columns:
                 val = row[field]
            else:
                     val = ""

            # Limpiar NA/NaN/None de pandas a string vacío
            try:
                if val is None or (hasattr(pd, "isna") and pd.isna(val)):
                    val = ""
            except (TypeError, ValueError):
                pass

            cell = ws.cell(ri, ci)
            cell.alignment = _aln()

            if field == "DIFERENCIA":
                try:
                    n = float(val)
                    cell.value         = n
                    cell.number_format = "#,##0.00"
                    neg = n < 0
                    cell.font = _font(bold=True, color="CC0000" if neg else "006100")
                    cell.fill = _fill("FFCCCC" if neg else "CCFFCC")
                except Exception:
                    cell.value = val; cell.font = _font(bold=True); cell.fill = _fill("FCE4D6")

            elif field in ("VAL_TARJETA", "VAL_EVENTO", "VAL_TRAMO", "VAL_CARRIL"):
                ok = str(val).strip().lower() in ("true", "verdadero", "1")
                cell.value = "VERDADERO" if ok else "FALSO"
                cell.font  = _font(bold=True, color="006100" if ok else "9C0006")
                cell.fill  = _fill("C6EFCE" if ok else "FFC7CE")

            elif field == "IMPORTE_VALUADO":
                try:
                    n = float(val); cell.value = n; cell.number_format = "#,##0"
                except Exception:
                    cell.value = val
                cell.font = _font(bold=True); cell.fill = _fill("FCE4D6")

            elif is_calc:
                cell.value = val
                cell.font  = _font(bold=True)
                cell.fill  = _fill("FFFF99")

            else:
                if field == "Hora Transaccion":
                    try:
                        s = str(val).strip().zfill(6)
                        cell.value = f"{s[:2]}:{s[2:4]}:{s[4:6]}" if s.isdigit() else str(val)
                    except Exception:
                        cell.value = val
                else:
                    cell.value = val
                cell.font  = _font()
                cell.fill  = _fill(SECTION_COLORS[sec]["dbg"] if odd else "FFFFFF")

            # Color por escenario
            if field == "ESCENARIO":
                escenario = str(val).strip()
                if escenario.startswith("Transaccion coincidente, sin diferencia"):
                    cell.font = _font(bold=True, color="3B6FCC")
                    cell.fill = _fill("FFFFFF")
                elif escenario.startswith("Transaccion coincidente, con diferencia"):
                    cell.font = _font(bold=True, color="FF0000")
                    cell.fill = _fill("FFFFFF")
                elif "ni en dia anterior o posterior" in escenario or "ni en DF de dia anterior" in escenario:
                    # Sobrante puro (DF o CCI): no se localizó en ningún día vecino
                    cell.font = _font(bold=True, color="FF0000")
                    cell.fill = _fill("FFFFFF")
                elif "Sobrante de DF" in escenario:
                    # localizada en CCI día anterior/posterior
                    cell.font = _font(bold=True, color="1F4E79")
                    cell.fill = _fill("DEEAF1")
                elif "Sobrante de CCI" in escenario:
                    # localizada en DF día anterior/siguiente
                    cell.font = _font(bold=True, color="1E4620")
                    cell.fill = _fill("E2EFDA")

    for ci, (_, label, _, _) in enumerate(ALL_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COLUMN_WIDTHS.get(label, DEFAULT_WIDTH)

    ws.freeze_panes = "A3"


# ── Hoja TARIFAS ──────────────────────────────────────────────────────────────

def _write_tarifas(ws):
    """
    Genera la hoja TARIFAS como tabla cruzada:
      Col A = CLASE, Col B = TRAMO 574, Col C = TRAMO 575, Col D = TRAMO 576
    Con secciones: Básicas, Larga Estancia (L01-L15), Paso Posterior (P01-P15).
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    TRAMOS = ["574", "575", "576"]

    # Colores por tramo 
    TRAMO_HDR_BG  = {"574": "C55A11",  "575": "375623",  "576": "1F4E79"}
    TRAMO_HDR_FG  = {"574": "FFFFFF",  "575": "FFFFFF",  "576": "FFFFFF"}
    TRAMO_DATA_BG = {"574": "FCE4D6",  "575": "E2EFDA",  "576": "DEEAF1"}

    # Secciones
    SECTIONS = [
        ("Tarifas Básicas", [
            "T01A","T01M",
            "T02B","T02C",
            "T03B","T03C",
            "T04B","T04C",
            "T05C","T06C",
            "T07C","T08C","T09C",
            "EEL","EEP",
        ]),
        ("Larga Estancia (T01L)", [f"T01L{i:02d}A" for i in range(1, 16)]),
        ("Paso Posterior (T09P)", [f"T09P{i:02d}C" for i in range(1, 16)]),
    ]

    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="595959")
    def border(top=None,bottom=None,left=None,right=None):
        return Border(top=top,bottom=bottom,left=left,right=right)

    def _set(ws, row, col, val="", bold=False, fg="000000", bg=None,
             size=9, wrap=False, num_fmt=None, b_top=None, b_bot=None, b_left=None, b_right=None):
        c = ws.cell(row, col, val)
        c.font = Font(name="Arial", bold=bold, color=fg, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if num_fmt:
            c.number_format = num_fmt
        c.border = Border(
            top=b_top or thin, bottom=b_bot or thin,
            left=b_left or thin, right=b_right or thin
        )
        return c

    ws.merge_cells("A1:D1")
    c = ws.cell(1, 1, "TARIFAS VIGENTES — PC 0231 · A PARTIR DEL 13/04/2026")
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill      = PatternFill("solid", fgColor="0B3D2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _set(ws, 2, 1, "CLASE", bold=True, fg="FFFFFF", bg="2E4057", size=10,
         b_left=thick, b_top=thick, b_bot=thick)
    for ci, tramo in enumerate(TRAMOS, 2):
        _set(ws, 2, ci, f"TRAMO {tramo}", bold=True,
             fg=TRAMO_HDR_FG[tramo], bg=TRAMO_HDR_BG[tramo], size=10,
             b_top=thick, b_bot=thick,
             b_right=thick if ci == 4 else thin)
    ws.row_dimensions[2].height = 22

    row = 3
    for sec_title, clases in SECTIONS:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row, 1, sec_title)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor="595959")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = Border(top=thick, bottom=thick, left=thick, right=thick)
        ws.row_dimensions[row].height = 16
        row += 1

        for idx, clase in enumerate(clases):
            is_even = idx % 2 == 0
            row_bg = "F5F5F5" if is_even else "FFFFFF"

            _set(ws, row, 1, clase, bold=False, fg="1a1a1a", bg=row_bg,
                 b_left=thick, b_right=thin)

            for ci, tramo in enumerate(TRAMOS, 2):
                key = f"{tramo}{clase}"
                val = TARIFAS.get(key, "—")
                is_num = isinstance(val, (int, float))
                _set(ws, row, ci,
                     val if is_num else val,
                     bold=is_num, fg="000000" if is_num else "999999",
                     bg=TRAMO_DATA_BG[tramo] if is_num else row_bg,
                     num_fmt="#,##0" if is_num else None,
                     b_right=thick if ci == 4 else thin)
            row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13

    ws.freeze_panes = "B3"

# ──  ──────────────────────────────────────────────────────────────


def _index_clr_otro(clr_otro: pd.DataFrame | None) -> dict:
    """
    Construye un índice INDICADOR_PISO -> dict con datos del CLR de un día contrario,
    para enriquecer las filas right_only del día HOY que se localizaron en el día contrario.
    Recibe el DataFrame de CLR ya pasado por build_indicators (el "merged" de ese día).
    """
    if clr_otro is None:
        return {}
    idx = {}
    # Solo nos interesan filas que tengan INDICADOR_PISO (lado CLR del merge)
    sub = clr_otro[clr_otro["INDICADOR_PISO"].notna() & (clr_otro["INDICADOR_PISO"].astype(str).str.strip() != "")]
    # Si hay duplicados, nos quedamos con el primero
    sub = sub.drop_duplicates(subset=["INDICADOR_PISO"], keep="first")
    for _, r in sub.iterrows():
        ind = str(r["INDICADOR_PISO"]).strip()
        idx[ind] = {
            "OTRO_Operador Carretero":    r.get("Operador Carretero"),
            "OTRO_Fecha Base":            r.get("Fecha Base"),
            "OTRO_Plaza de Cobro":        r.get("Plaza de Cobro"),
            "OTRO_Numero de Transaccion": r.get("Numero de Transaccion"),
            "OTRO_Estado Transaccion":    r.get("Estado Transaccion"),
            "OTRO_Descripcion":           r.get("Descripcion"),
            "OTRO_Fecha Transaccion":     r.get("Fecha Transaccion"),
            "OTRO_Hora Transaccion":      r.get("Hora Transaccion"),
            "OTRO_Numero Tarjeta":        r.get("Numero Tarjeta"),
            "OTRO_Estado Tarjeta":        r.get("Estado Tarjeta"),
            "OTRO_Carril":                r.get("Carril"),
            "OTRO_Evento":                r.get("Evento"),
            "OTRO_Clase":                 r.get("Clase"),
            "OTRO_Tipo":                  r.get("Tipo"),
            "OTRO_Tramo":                 r.get("Tramo"),
            "OTRO_Forma de pago Capufe":  r.get("Forma de pago Capufe"),
            "OTRO_CONVERSION_HORA":       r.get("CONVERSION_HORA"),
            "OTRO_INDICADOR_PISO":        r.get("INDICADOR_PISO"),
            "OTRO_INDICADOR_TARIFA":      r.get("INDICADOR_TARIFA"),
            "OTRO_IMPORTE_VALUADO":       r.get("IMPORTE_VALUADO"),
        }
    return idx


def _build_scenarios(df_hoy: pd.DataFrame, merged_ant: pd.DataFrame | None, merged_pos: pd.DataFrame | None):
    """
    Lógica completa replicando el manual:

    1) BOTH (cruzan en HOY)
       → "Transaccion coincidente, sin/con diferencia de importe"

    2) LEFT_ONLY (sobrante de DF de HOY): buscar el INDICADOR_PISO en el INDICADOR_CCI
       de los días vecinos.
       - en CCI ANTERIOR  → "Sobrante de DF, localizada en CCI de día anterior, sin/con diferencia"
       - en CCI POSTERIOR → "Sobrante de DF, localizada en CCI de día posterior, sin/con diferencia"
       - ninguno          → "Sobrante de DF y no en CCI, ni en dia anterior o posterior"

    3) RIGHT_ONLY (sobrante de CCI de HOY): buscar el INDICADOR_CCI en el INDICADOR_PISO
       de los días vecinos.
       - en DF ANTERIOR   → "Sobrante de CCI, localizada en DF de día anterior, sin/con diferencia"
       - en DF POSTERIOR  → "Sobrante de CCI, localizada en DF de día siguiente, sin/con diferencia"
       - ninguno          → "Sobrante de CCI y no en DF de dia anterior o dia siguiente"

       Cuando un right_only se localiza en otro día, se llenan las columnas OTRO_* con los
       datos del DF cruzado y la diferencia se calcula como
       |IMPORTE_VALUADO_OTRO - IMPORTE_TOTAL_CCI|.
    """
    df = df_hoy.copy()

    # ── columnas RESULTADO base ─────────────────────────────────────────────
    df["DF_DIA"] = ""
    df["CCI_DIA"] = ""
    df["DIF_IMPORTE"] = False
    df["ESCENARIO"] = "SIN_CLASIFICAR"

    # Asegurar que existan las columnas OTRO_* aunque queden vacías
    for col, _, _, _ in DF_OTRO_COLS:
        if col not in df.columns:
            df[col] = pd.NA

    # ── Índices auxiliares ──────────────────────────────────────────────────
    # set de INDICADOR_CCI presentes en cada día vecino (lado CCI)
    nums_cci_ant = set()
    nums_cci_pos = set()
    if merged_ant is not None and "INDICADOR_CCI" in merged_ant.columns:
        nums_cci_ant = set(merged_ant["INDICADOR_CCI"].dropna().astype(str).str.strip())
        nums_cci_ant.discard("")
    if merged_pos is not None and "INDICADOR_CCI" in merged_pos.columns:
        nums_cci_pos = set(merged_pos["INDICADOR_CCI"].dropna().astype(str).str.strip())
        nums_cci_pos.discard("")

    # set de INDICADOR_PISO presentes en cada día vecino (lado CLR/DF)
    nums_df_ant = set()
    nums_df_pos = set()
    if merged_ant is not None and "INDICADOR_PISO" in merged_ant.columns:
        nums_df_ant = set(merged_ant["INDICADOR_PISO"].dropna().astype(str).str.strip())
        nums_df_ant.discard("")
    if merged_pos is not None and "INDICADOR_PISO" in merged_pos.columns:
        nums_df_pos = set(merged_pos["INDICADOR_PISO"].dropna().astype(str).str.strip())
        nums_df_pos.discard("")

    # Índices INDICADOR_PISO -> datos del DF de cada día vecino (para llenar OTRO_*)
    idx_df_ant = _index_clr_otro(merged_ant)
    idx_df_pos = _index_clr_otro(merged_pos)

    # Índices INDICADOR_CCI -> IMPORTE_TOTAL del CCI de cada día vecino
    def _idx_cci_importe(merged_otro):
        if merged_otro is None or "INDICADOR_CCI" not in merged_otro.columns:
            return {}
        sub = merged_otro[merged_otro["INDICADOR_CCI"].notna()].drop_duplicates(subset=["INDICADOR_CCI"], keep="first")
        return dict(zip(sub["INDICADOR_CCI"].astype(str).str.strip(), sub.get("IMPORTE_TOTAL", pd.Series(dtype=float))))
    imp_cci_ant = _idx_cci_importe(merged_ant)
    imp_cci_pos = _idx_cci_importe(merged_pos)

    # Helper: convertir a float seguro
    def _f(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)): return None
            s = str(v).replace(",", "").strip()
            if s == "" or s.lower() == "nan": return None
            return float(s)
        except Exception:
            return None

    # =========================================================================
    # CASO 1: BOTH (coincidentes del mismo día)
    # =========================================================================
    mask_both = df["_merge"] == "both"
    # Diferencia para coincidentes: usa la columna DIFERENCIA ya calculada
    if "DIFERENCIA" in df.columns:
        dif_num = df.loc[mask_both, "DIFERENCIA"].apply(_f).fillna(0)
        df.loc[mask_both, "DIF_IMPORTE"] = dif_num.abs().round(2) > 0
    df.loc[mask_both & (~df["DIF_IMPORTE"]), "ESCENARIO"] = "Transaccion coincidente, sin diferencia de importe"
    df.loc[mask_both & ( df["DIF_IMPORTE"]), "ESCENARIO"] = "Transaccion coincidente, con diferencia de importe"
    df.loc[mask_both, "DF_DIA"]  = "HOY"
    df.loc[mask_both, "CCI_DIA"] = "HOY"

    # =========================================================================
    # CASO 2: LEFT_ONLY (sobrantes de DF del día HOY)
    # =========================================================================
    mask_left = df["_merge"] == "left_only"
    df.loc[mask_left, "DF_DIA"]  = "HOY"
    df.loc[mask_left, "CCI_DIA"] = ""

    # 2a) localizada en CCI día ANTERIOR
    mask_left_en_ant = mask_left & df["INDICADOR_PISO"].astype(str).str.strip().isin(nums_cci_ant)
    for idx in df[mask_left_en_ant].index:
        ind = str(df.at[idx, "INDICADOR_PISO"]).strip()
        imp_val = _f(df.at[idx, "IMPORTE_VALUADO"])
        imp_cci = _f(imp_cci_ant.get(ind))
        if imp_val is None or imp_cci is None:
            con_dif = False
            diff = imp_val if imp_val is not None else 0
        else:
            diff = round(imp_val - imp_cci, 2)
            con_dif = abs(diff) > 0
        sufijo = "con diferencia de importe" if con_dif else "sin diferencia de importe"
        df.at[idx, "ESCENARIO"]   = f"Sobrante de DF, localizada en CCI de día anterior, {sufijo}"
        df.at[idx, "CCI_DIA"]     = "ANTERIOR"
        df.at[idx, "DIF_IMPORTE"] = con_dif
        df.at[idx, "DIFERENCIA"]  = diff

    # 2b) localizada en CCI día POSTERIOR (entre los que aún no se ubicaron)
    mask_left_pendiente = mask_left & (df["ESCENARIO"] == "SIN_CLASIFICAR")
    mask_left_en_pos = mask_left_pendiente & df["INDICADOR_PISO"].astype(str).str.strip().isin(nums_cci_pos)
    for idx in df[mask_left_en_pos].index:
        ind = str(df.at[idx, "INDICADOR_PISO"]).strip()
        imp_val = _f(df.at[idx, "IMPORTE_VALUADO"])
        imp_cci = _f(imp_cci_pos.get(ind))
        if imp_val is None or imp_cci is None:
            con_dif = False
            diff = imp_val if imp_val is not None else 0
        else:
            diff = round(imp_val - imp_cci, 2)
            con_dif = abs(diff) > 0
        sufijo = "con diferencia de importe" if con_dif else "sin diferencia de importe"
        df.at[idx, "ESCENARIO"]   = f"Sobrante de DF, localizada en CCI de día posterior, {sufijo}"
        df.at[idx, "CCI_DIA"]     = "POSTERIOR"
        df.at[idx, "DIF_IMPORTE"] = con_dif
        df.at[idx, "DIFERENCIA"]  = diff

    # 2c) sobrante puro de DF
    mask_left_puro = mask_left & (df["ESCENARIO"] == "SIN_CLASIFICAR")
    df.loc[mask_left_puro, "ESCENARIO"] = "Sobrante de DF y no en CCI, ni en dia anterior o posterior"

    # =========================================================================
    # CASO 3: RIGHT_ONLY (sobrantes de CCI del día HOY)
    # =========================================================================
    mask_right = df["_merge"] == "right_only"
    df.loc[mask_right, "CCI_DIA"] = "HOY"
    df.loc[mask_right, "DF_DIA"]  = ""

    # 3a) localizada en DF día ANTERIOR
    mask_right_en_ant = mask_right & df["INDICADOR_CCI"].astype(str).str.strip().isin(nums_df_ant)
    for idx in df[mask_right_en_ant].index:
        ind = str(df.at[idx, "INDICADOR_CCI"]).strip()
        otro = idx_df_ant.get(ind, {})
        imp_otro = _f(otro.get("OTRO_IMPORTE_VALUADO"))
        imp_cci  = _f(df.at[idx, "IMPORTE_TOTAL"])
        if imp_otro is None or imp_cci is None:
            con_dif = False
            diff = -imp_cci if imp_cci is not None else 0
        else:
            diff = round(imp_otro - imp_cci, 2)
            con_dif = abs(diff) > 0
        sufijo = "con diferencia de importe" if con_dif else "sin diferencia de importe"
        df.at[idx, "ESCENARIO"]   = f"Sobrante de CCI, localizada en DF de día anterior, {sufijo}"
        df.at[idx, "DF_DIA"]      = "ANTERIOR"
        df.at[idx, "DIF_IMPORTE"] = con_dif
        df.at[idx, "DIFERENCIA"]  = diff
        for k, v in otro.items():
            df.at[idx, k] = v

    # 3b) localizada en DF día SIGUIENTE/POSTERIOR
    mask_right_pendiente = mask_right & (df["ESCENARIO"] == "SIN_CLASIFICAR")
    mask_right_en_pos = mask_right_pendiente & df["INDICADOR_CCI"].astype(str).str.strip().isin(nums_df_pos)
    for idx in df[mask_right_en_pos].index:
        ind = str(df.at[idx, "INDICADOR_CCI"]).strip()
        otro = idx_df_pos.get(ind, {})
        imp_otro = _f(otro.get("OTRO_IMPORTE_VALUADO"))
        imp_cci  = _f(df.at[idx, "IMPORTE_TOTAL"])
        if imp_otro is None or imp_cci is None:
            con_dif = False
            diff = -imp_cci if imp_cci is not None else 0
        else:
            diff = round(imp_otro - imp_cci, 2)
            con_dif = abs(diff) > 0
        sufijo = "con diferencia de importe" if con_dif else "sin diferencia de importe"
        df.at[idx, "ESCENARIO"]   = f"Sobrante de CCI, localizada en DF de día siguiente, {sufijo}"
        df.at[idx, "DF_DIA"]      = "POSTERIOR"
        df.at[idx, "DIF_IMPORTE"] = con_dif
        df.at[idx, "DIFERENCIA"]  = diff
        for k, v in otro.items():
            df.at[idx, k] = v

    # 3c) sobrante puro de CCI
    mask_right_puro = mask_right & (df["ESCENARIO"] == "SIN_CLASIFICAR")
    df.loc[mask_right_puro, "ESCENARIO"] = "Sobrante de CCI y no en DF de dia anterior o dia siguiente"

    # =========================================================================
    # AJUSTES: reglas explícitas por escenario
    # =========================================================================
    # Por default: van a AJUSTES los que tengan DIF_IMPORTE=True
    # Excepción 1: "Sobrante de DF, localizada en CCI de día anterior, con diferencia"
    #              NO va a ajustes aunque tenga diferencia.
    # Excepción 2: "Sobrante de DF y no en CCI, ni en dia anterior o posterior"
    #              SIEMPRE va a ajustes, aunque no tenga diferencia.
    EXCLUIR_AJUSTES = {
        "Sobrante de DF, localizada en CCI de día anterior, con diferencia de importe",
    }
    SIEMPRE_AJUSTES = {
        "Sobrante de DF y no en CCI, ni en dia anterior o posterior",
    }
    mask_ajustes = (
        ((df["DIF_IMPORTE"] == True) & (~df["ESCENARIO"].isin(EXCLUIR_AJUSTES)))
        | df["ESCENARIO"].isin(SIEMPRE_AJUSTES)
    )
    ajustes = df[mask_ajustes].copy()

    orden = {
        "Transaccion coincidente, sin diferencia de importe": 0,
        "Transaccion coincidente, con diferencia de importe": 1,
        "Sobrante de DF, localizada en CCI de día anterior, sin diferencia de importe":  2,
        "Sobrante de DF, localizada en CCI de día anterior, con diferencia de importe":  3,
        "Sobrante de DF, localizada en CCI de día posterior, sin diferencia de importe": 4,
        "Sobrante de DF, localizada en CCI de día posterior, con diferencia de importe": 5,
        "Sobrante de DF y no en CCI, ni en dia anterior o posterior":                    6,
        "Sobrante de CCI, localizada en DF de día anterior, sin diferencia de importe":  7,
        "Sobrante de CCI, localizada en DF de día anterior, con diferencia de importe":  8,
        "Sobrante de CCI, localizada en DF de día siguiente, sin diferencia de importe": 9,
        "Sobrante de CCI, localizada en DF de día siguiente, con diferencia de importe": 10,
        "Sobrante de CCI y no en DF de dia anterior o dia siguiente":                    11,
        "SIN_CLASIFICAR":                                                                99,
    }
    df["_orden"] = df["ESCENARIO"].map(orden).fillna(99)
    df = df.sort_values("_orden").drop(columns=["_orden"]).reset_index(drop=True)

    return df, ajustes


# ── Función pública
def _split_and_promote_ajustes(ajustes: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Divide los ajustes en dos bloques para mejor visualización:
      - bloque "HOY": filas cuyos datos del DF corresponden al día que se está conciliando
        (Transaccion coincidente con diferencia, Sobrante de DF y no en CCI, etc.)
      - bloque "DIA SIGUIENTE": filas "Sobrante de CCI, localizada en DF de día siguiente,
        con diferencia de importe" donde los datos reales del DF están en las columnas OTRO_*.
        En este bloque, se promueven las columnas OTRO_* a las columnas DF principales para
        que se vean alineadas con el resto.
    """
    if ajustes is None or len(ajustes) == 0:
        return ajustes, ajustes.iloc[0:0] if ajustes is not None else ajustes

    mask_siguiente = ajustes["ESCENARIO"].astype(str).str.contains(
        "Sobrante de CCI, localizada en DF de día siguiente", na=False
    )
    bloque_hoy = ajustes[~mask_siguiente].copy()
    bloque_sig = ajustes[mask_siguiente].copy()

    # Promover OTRO_* a columnas DF principales en bloque_sig
    PROMOCION = {
        "OTRO_Operador Carretero":    "Operador Carretero",
        "OTRO_Fecha Base":            "Fecha Base",
        "OTRO_Plaza de Cobro":        "Plaza de Cobro",
        "OTRO_Numero de Transaccion": "Numero de Transaccion",
        "OTRO_Estado Transaccion":    "Estado Transaccion",
        "OTRO_Descripcion":           "Descripcion",
        "OTRO_Fecha Transaccion":     "Fecha Transaccion",
        "OTRO_Hora Transaccion":      "Hora Transaccion",
        "OTRO_Numero Tarjeta":        "Numero Tarjeta",
        "OTRO_Estado Tarjeta":        "Estado Tarjeta",
        "OTRO_Carril":                "Carril",
        "OTRO_Evento":                "Evento",
        "OTRO_Clase":                 "Clase",
        "OTRO_Tipo":                  "Tipo",
        "OTRO_Tramo":                 "Tramo",
        "OTRO_Forma de pago Capufe":  "Forma de pago Capufe",
        "OTRO_CONVERSION_HORA":       "CONVERSION_HORA",
        "OTRO_INDICADOR_PISO":        "INDICADOR_PISO",
        "OTRO_INDICADOR_TARIFA":      "INDICADOR_TARIFA",
        "OTRO_IMPORTE_VALUADO":       "IMPORTE_VALUADO",
    }
    if len(bloque_sig) > 0:
        for src, dst in PROMOCION.items():
            if src in bloque_sig.columns and dst in bloque_sig.columns:
                bloque_sig[dst] = bloque_sig[src]
        # Limpiar las columnas OTRO_* del bloque siguiente (ya se promovieron)
        for src in PROMOCION.keys():
            if src in bloque_sig.columns:
                bloque_sig[src] = pd.NA

    return bloque_hoy, bloque_sig


def _write_ajustes(ws, ajustes: pd.DataFrame):
    """
    Escribe la hoja AJUSTES en dos bloques con un separador entre ellos:
      - Bloque 1: filas del día HOY (formato normal)
      - Bloque 2: filas "Sobrante de CCI, día siguiente" (fondo gris, datos del DF del día +1)
    """
    bloque_hoy, bloque_sig = _split_and_promote_ajustes(ajustes)

    # Si no hay filas "día siguiente", se escribe como una hoja normal
    if len(bloque_sig) == 0:
        _write_conciliacion(ws, ajustes)
        return

    # Escribir bloque 1 (HOY) usando la función normal
    _write_conciliacion(ws, bloque_hoy)

    # Localizar la siguiente fila libre (después del bloque 1)
    next_row = ws.max_row + 2  # un renglón en blanco como separador

    # Header del bloque 2 (igual al header normal, repetido)
    # Fila 1: secciones (DF, VALUACION, CCI, DF_OTRO)
    # Fila 2: nombres de columnas
    header_sec_row = next_row
    header_col_row = next_row + 1

    # Repetir cabeceras como en _write_conciliacion
    prev_sec = None
    sec_start = 1
    for ci, (_, label, sec, _) in enumerate(ALL_COLS, 1):
        if sec != prev_sec and prev_sec is not None:
            ws.merge_cells(start_row=header_sec_row, start_column=sec_start,
                           end_row=header_sec_row, end_column=ci - 1)
            c = ws.cell(header_sec_row, sec_start)
            c.value     = prev_sec
            c.font      = _font(True, SECTION_COLORS[prev_sec]["hfg"], 12)
            c.fill      = _fill(SECTION_COLORS[prev_sec]["hbg"])
            c.alignment = _aln()
            sec_start = ci
        prev_sec = sec
    # último bloque
    ws.merge_cells(start_row=header_sec_row, start_column=sec_start,
                   end_row=header_sec_row, end_column=len(ALL_COLS))
    c = ws.cell(header_sec_row, sec_start)
    c.value     = prev_sec
    c.font      = _font(True, SECTION_COLORS[prev_sec]["hfg"], 12)
    c.fill      = _fill(SECTION_COLORS[prev_sec]["hbg"])
    c.alignment = _aln()

    # Cabecera de columnas
    for ci, (_, label, sec, _) in enumerate(ALL_COLS, 1):
        cell = ws.cell(header_col_row, ci)
        cell.value     = label
        cell.font      = _font(True, "000000", 9)
        cell.fill      = _fill(SECTION_COLORS[sec]["dbg"])
        cell.alignment = _aln(wrap=True)
    ws.row_dimensions[header_col_row].height = 34

    # Datos del bloque 2 con fondo gris
    GRIS_FILL = "BFBFBF"
    GRIS_FILL_ALT = "D9D9D9"
    data_start = header_col_row + 1
    for ri, (_, row) in enumerate(bloque_sig.iterrows(), data_start):
        odd = (ri % 2 == 0)
        for ci, (field, _, sec, is_calc) in enumerate(ALL_COLS, 1):
            val = row[field] if field in bloque_sig.columns else ""
            try:
                if val is None or (hasattr(pd, "isna") and pd.isna(val)):
                    val = ""
            except (TypeError, ValueError):
                pass

            cell = ws.cell(ri, ci)
            cell.alignment = _aln()

            # Formateo del valor según el tipo de columna (igual que _write_conciliacion)
            if field == "DIFERENCIA":
                try:
                    n = float(val); cell.value = n; cell.number_format = "#,##0.00"
                except Exception:
                    cell.value = val
            elif field in ("VAL_TARJETA", "VAL_EVENTO", "VAL_TRAMO", "VAL_CARRIL"):
                ok = str(val).strip().lower() in ("true", "verdadero", "1")
                cell.value = "VERDADERO" if ok else "FALSO"
            elif field == "IMPORTE_VALUADO":
                try:
                    n = float(val); cell.value = n; cell.number_format = "#,##0"
                except Exception:
                    cell.value = val
            elif field == "Hora Transaccion":
                try:
                    s = str(val).strip().zfill(6)
                    cell.value = f"{s[:2]}:{s[2:4]}:{s[4:6]}" if s.isdigit() else str(val)
                except Exception:
                    cell.value = val
            else:
                cell.value = val

            # Fondo gris uniforme + texto en negro
            cell.font = _font(bold=(field == "ESCENARIO"), color="000000")
            cell.fill = _fill(GRIS_FILL_ALT if odd else GRIS_FILL)

            # Resaltar el ESCENARIO
            if field == "ESCENARIO":
                cell.font = _font(bold=True, color="1E4620")


# ── Función pública
def generate_excel(df_hoy: pd.DataFrame, merged_ant=None, merged_pos=None) -> io.BytesIO:
    wb = Workbook()

    conciliacion, ajustes = _build_scenarios(df_hoy, merged_ant, merged_pos)

    ws_main = wb.active
    ws_main.title = "CONCILIACION"
    _write_conciliacion(ws_main, conciliacion)

    ws_aj = wb.create_sheet("AJUSTES")
    _write_ajustes(ws_aj, ajustes)

    ws_tar = wb.create_sheet("TARIFAS")
    _write_tarifas(ws_tar)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf